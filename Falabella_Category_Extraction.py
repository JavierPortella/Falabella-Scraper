# Librerías a importar
from concurrent.futures import as_completed, ThreadPoolExecutor
from datetime import datetime, timedelta
from json import JSONDecodeError
from logging import (
    Formatter,
    getLogger,
    FileHandler,
    INFO,
    shutdown,
    StreamHandler,
)
from os import getcwd, makedirs, path
from re import search, sub
from sys import stdout
from time import localtime, strftime, time
from traceback import TracebackException

from openpyxl import load_workbook, Workbook
from pandas import concat, DataFrame, read_csv
from requests import get
from selenium.common.exceptions import TimeoutException, ElementNotInteractableException
from seleniumwire.webdriver import Chrome, ChromeOptions
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from urllib.parse import quote_plus
from webdriver_manager.chrome import ChromeDriverManager

# Constantes usadas en el script
CURRENT_DATE = datetime.now().date()
ROOT_PATH = getcwd()
URL_FALABELLA = "https://www.falabella.com.pe/falabella-pe"
DATA_FILENAME = "falabella_category"
DATA_FOLDER = "Data"
METADATA_FILENAME = "Metadata.xlsx"
METADATA_SHEET_NAME = "Categorias"
DATA_DICT_FILENAME = "category_dictionary.csv"
DATA_DICT_HEADERS = ["Link_subcat", "Name", "Link_cat"]
API_HEADERS = {
    "accept": "*/*",
    "accept-language": "es,es-ES;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
    "content-type": "application/json",
    "sec-ch-ua": '"Chromium";v="110", "Not A(Brand";v="24", "Microsoft Edge";v="110"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-origin",
    "x-device-type": "desktop",
    "Referer": "https://www.falabella.com.pe/",
    "Referrer-Policy": "strict-origin-when-cross-origin",
}
API_URL = "https://www.falabella.com.pe/s/browse/v1/listing/pe?=&\
{0}&page=1&categoryId={1}&categoryName={2}&pgid=2&pid=799c102f-9b4c-44be-a421-23e366a63b82\
&zones=912_LIMA_2%2COLVAA_81%2CLIMA_URB1_DIRECTO%2CURBANO_83%2CIBIS_19%2C912_LIMA_1%2C150101%2CPERF_TEST%2C150000"
THREAD = ThreadPoolExecutor()
LOGGER = getLogger(__name__)


class Metadata:
    """Representa a la información generada durante la ejecución del scraper

    Attributes:
        start_time (float): Hora de inicio de la ejecución del scraper en segundos
        execution_date (str): Fecha de extracción de las categorias en formato %d/%m/%Y
        start_hour (str): Hora de inicio de la ejecución del scraper en formato %H:%M:%S
        end_hour (str): Hora de término de la ejecución del scraper en formato %H:%M:%S
        quantity (int): Cantidad de categorías extraídas de la página de saga falabella
        time_execution (str): Tiempo de ejecución del scraper en formato %d days, %H:%M:%S
        category_per_min (float): Cantidad de categorías que puede extraer el scraper en un minuto
        num_errors (int): Cantidad de errores ocurridos durante la ejecución del scraper
    """

    def __init__(self):
        """Genera todos los atributos para una instancia de la clase Metadata"""
        self._start_time = time()
        self._execution_date = CURRENT_DATE.strftime("%d/%m/%Y")
        self._start_hour = strftime("%H:%M:%S", localtime(self._start_time))
        self._end_hour = 0
        self._quantity = 0
        self._time_execution = 0
        self._category_per_min = 0
        self._num_errors = 0
        LOGGER.info(f"Hora de inicio: {self._start_hour}")

    @property
    def execution_date(self):
        """Retorna el valor actual del atributo execution_date"""
        return self._execution_date

    @property
    def num_errors(self):
        """Retorna el valor actual o actualiza el valor del atributo num_errors"""
        return self._num_errors

    @property
    def quantity(self):
        """Retorna el valor actual o actualiza el valor del atributo quantity"""
        return self._quantity

    @quantity.setter
    def num_errors(self, num_errors):
        self._num_errors = num_errors

    @quantity.setter
    def quantity(self, quantity):
        self._quantity = quantity

    def set_param_final(self):
        """Registra los atributos restantes de la clase MetaData"""
        end = time()
        self._end_hour = strftime("%H:%M:%S", localtime(end))
        total = end - self._start_time
        self._time_execution = str(timedelta(seconds=total)).split(".")[0]
        self._category_per_min = round(self._quantity / (total / 60), 2)
        LOGGER.info(f"Se halló {self._num_errors} errores")
        LOGGER.info(f"Categorías Extraídas: {self._quantity}")
        LOGGER.info(f"Hora Fin: {self._end_hour}")


class Error(TracebackException):
    """Extiende la clase TracebackException para el manejo del traceback"""

    def __init__(self, error) -> None:
        """Genera todos los atributos para una instancia de la clase Error

        Args:
            error (Exception): Error ocurrido durante la ejecución del scraper
        """
        super().__init__(type(error), error, error.__traceback__)

    def imprimir_error(self):
        """Imprime toda la información del error por consola"""
        LOGGER.error("Ha ocurrido un error:")
        for line in self.format(chain=True):
            LOGGER.error(line)


class WebDriver(Chrome):
    """Extiende la clase Chrome WebDriver para añadirle un tiempo de espera

    Attributes:
        wait (selenium.webdriver.support.wait.WebDriverWait): Atributo que maneja el tiempo máximo de espera usado por el navegador para buscar los elementos
    """

    def __init__(self, timeout=7):
        """Inicializa una instancia de Chrome WebDriver

        Args:
            timeout (int, optional): Tiempo máximo de espera en segundos. Defaults to 7.
        """
        chrome_options = ChromeOptions()
        prefs = {
            "profile.default_content_setting_values.notifications": 2,
            "profile.managed_default_content_settings.popups": 2,
        }
        chrome_options.add_experimental_option("prefs", prefs)
        chrome_options.add_experimental_option(
            "excludeSwitches", ["enable-logging"]
        )  # Suprimir los mensajes de consola
        seleniumwire_options = {"disable_capture": True}  # No guardar ningún request
        super().__init__(
            options=chrome_options,
            seleniumwire_options=seleniumwire_options,
            service=Service(ChromeDriverManager().install()),
        )
        self._wait = WebDriverWait(self, timeout)
        self.maximize_window()

    def get_element(self, method, message=""):
        """Función que busca uno o varios elementos ubicados en la página web y los retorna si los encuentra dentro de un tiempo establecido

        Args:
            method (Any): Llamada a una función de búsqueda de uno o varios elementos ubicados en la página web
            message (str, optional): Mensaje a mostrar en caso la búsqueda falle. Defaults to "".

        Returns:
            Any: El resultado devuelto al llamar a la función de búsqueda
        """
        return self._wait.until(method, message)


class ScraperFalabellaCategory:
    """Representa a un bot para hacer web scraping en saga falabella

    Attributes:
        metadata (Metadata): Objeto de la clase Metadata que maneja información generada durante la ejecución del scraper
        df_category (pandas.core.frame.DataFrame): Objeto de la clase DataFrame que maneja información de las categorías extraídas por el scraper
        df_dict_category (Dataset): Objeto de la clase Data que funciona como diccionario para mapear las categorías de saga falabella
        df_dict_category_filename (str): Nombre del archivo que contiene el diccionario de datos para mapear las categorías de saga falabella
        driver (WebDriver): Objeto de la clase WebDriver que maneja un navegador para hacer web scraping
    """

    def __init__(self, dict_filename):
        """Genera todos los atributos para una instancia de la clase ScraperFalabellaCategory

        Args:
            dict_filename (str): Nombre del archivo que va a ser usado como diccionario de datos
        """
        self._metadata = Metadata()
        self._df_category = DataFrame()
        # Comprobando si el diccionario para las categorías ya ha sido creado
        if path.isfile(dict_filename):
            self._df_dict_category = DataFrame(
                read_csv(dict_filename, names=DATA_DICT_HEADERS, encoding="utf-8-sig")
            )
            LOGGER.info(
                "El diccionario de categorías se ha definido satisfactoriamente",
            )
        else:
            self._df_dict_category = None
            LOGGER.info(
                "El diccionario de categorías no se va a utilizar por ser la primera ejecución",
            )
        self._df_dict_category_filename = dict_filename
        self._driver = WebDriver()

    def close_popups(self):
        """Cierra todas las ventanas emergentes que nos muestra la página principal de Saga Falabella"""
        marks = [
            (By.CLASS_NAME, "dy-lb-close"),
            (By.ID, "testId-accept-cookies-btn"),
            (By.CLASS_NAME, "airship-btn airship-btn-deny"),
            (By.CLASS_NAME, "dy-lb-close"),
        ]
        for selector, xpath in marks:
            try:
                self._driver.get_element(
                    EC.element_to_be_clickable((selector, xpath))
                ).click()
            except:
                pass

    def is_permanent_category(self, text):
        """Comprueba si la categoría mostrada por el menú es creada solo por temporadas o es permanente

        Args:
            text (str): Nombre de la categoría

        Returns:
            bool: Booleano
        """
        return text.split("\n")[-1] not in ["NUEVO", "Emprendedores", "SALE"]

    def is_url_category(self, url):
        """Comprueba si el link no pertenece a una categoría de falabella

        Args:
            url (str): Enlace web

        Returns:
            bool: Booleano que indica si el link pertenece o no a una categoría de falabella
        """
        return url.find("category") != -1

    def get_menu_links(self):
        """Función que navega por el menú de Saga Falabella y extrae todos los links mostrados por dicha interfaz

        Returns:
            list: Lista de enlaces
        """

        menu_links = []
        LOGGER.info("Accediendo al menú principal de saga falabella")
        self._driver.get_element(
            EC.element_to_be_clickable(
                (By.CLASS_NAME, "MarketplaceHamburgerBtn-module_hamburgerBtn__61t-r")
            )
        ).click()

        LOGGER.info(
            "Registrando la lista de categorías que nos muestra el menú principal"
        )
        category_list = self._driver.get_element(
            EC.visibility_of_all_elements_located(
                (
                    By.XPATH,
                    "//div[@class='FirstLevelCategories-module_categories__x82VK']",
                )
            )
        )

        LOGGER.info("Filtrando categorías que son creadas temporalmente")
        category_list = list(
            filter(
                lambda x: x,
                list(
                    THREAD.map(
                        lambda x: x if self.is_permanent_category(x.text) else None,
                        category_list,
                    )
                ),
            )
        )

        LOGGER.info("Navegando por el menú principal")
        for category in category_list:
            try:
                category.click()
                subcategory_list = self._driver.get_element(
                    EC.presence_of_all_elements_located(
                        (
                            By.XPATH,
                            "//li[@class='SecondLevelCategories-module_thirdLevelCategory__2ZQFF']/a",
                        )
                    )
                )
                menu_links += list(
                    THREAD.map(
                        lambda x: sub(r"\?.+", "", x.get_attribute("href")),
                        subcategory_list,
                    )
                )

            except TimeoutException:
                LOGGER.error(
                    "Tiempo agotado para recuperar las subcategorías mostradas por el menú principal de saga falabella",
                )

        LOGGER.info("Filtrando links duplicados")
        menu_links = list(set(menu_links))

        LOGGER.info("Filtrando los links que no pertenecen a una categoría")
        menu_links = list(
            filter(
                lambda x: x,
                list(
                    THREAD.map(
                        lambda x: x if self.is_url_category(x) else None,
                        menu_links,
                    )
                ),
            )
        )
        LOGGER.info(f"Se han extraído satisfactoriamente {len(menu_links)} links")
        return menu_links

    def get_category_info(self, subcategory_links):
        """Retorna un conjunto de datos que contiene toda la información de las categorías principales de saga falabella

        Args:
            subcategory_links (list): Lista de links correspondientes a las subcategorías encontradas en el menú de Saga Falabella

        Returns:
            pandas.core.frame.DataFrame: Instancia de la clase DataFrame
        """
        # Diccionario de datos que almacena la información de las categorías de saga falabella
        category_info_link = {}
        # Diccionario de datos que almacena las categorías de los links recorridos
        category_dict_info = {
            "Link_subcat": [],
            "Name": [],
            "Link_cat": [],
        }

        LOGGER.info(
            "Recopilando los links de las categorías principales a partir de los links presentados por el menú de Saga Falabella",
        )
        # Comprobando si el diccionario de links recorridos ha sido definido
        if self._df_dict_category is not None:
            LOGGER.info(
                "Usando el diccionario de datos para encontrar las categorías principales",
            )
            temp_subcat_links = []
            # Recorriendo los links de cada subcategoría
            for link in subcategory_links:
                results = self._df_dict_category[
                    self._df_dict_category[DATA_DICT_HEADERS[0]] == link
                ].values.tolist()
                # Comprobando que existan resultados
                if len(results) > 0:
                    # Guardando la información de la primera coincidencia
                    _, name, url_cat = results[0]
                    category_info_link[name] = url_cat
                else:
                    # Guardar los links que aún faltan recorrer
                    temp_subcat_links.append(link)

            LOGGER.info(
                f"El diccionario de datos ha mapeado {len(subcategory_links) - len(temp_subcat_links)} links de las subcategorías"
            )
            subcategory_links = temp_subcat_links
            del temp_subcat_links

        LOGGER.info(f"Hay {len(subcategory_links)} links que faltan recorrer")
        try:
            self._driver.get(subcategory_links[0])
            self._driver.get_element(
                EC.element_to_be_clickable((By.ID, "testId-modal-close"))
            ).click()
        except:
            pass

        # Recorriendo los links faltantes
        for link in subcategory_links:
            self._driver.get(link)
            # Comprobando que el link no te rediriga a otra página
            current_link = self._driver.execute_script("return document.URL")
            if not self.is_url_category(current_link):
                LOGGER.info(
                    f"No se va a extraer categorías del link {link}, pues te redirige a otro link: {current_link}",
                )
                continue

            # Navegar hasta la categoría padre de la subcategoría
            is_not_root_path = True
            while is_not_root_path:
                try:
                    self._driver.get_element(
                        EC.presence_of_element_located(
                            (By.XPATH, "//a[@class='jsx-2883309125 l1category']")
                        )
                    ).click()

                except ElementNotInteractableException:
                    LOGGER.info("Se ha conseguido llegar hasta la categoría principal")
                    is_not_root_path = False

            url_cat = self._driver.execute_script("return document.URL")
            name_cat = self._driver.get_element(
                EC.presence_of_element_located(
                    (By.XPATH, "//h1[@class='jsx-2883309125 l2category']")
                ),
            ).text

            # Guardando las nuevas incidencias al diccionario de categorías
            category_dict_info["Link_subcat"].append(link)
            category_dict_info["Name"].append(name_cat)
            category_dict_info["Link_cat"].append(url_cat)
            category_info_link[name_cat] = url_cat
            LOGGER.info(f"Categoría Obtenida: {name_cat}")

        LOGGER.info("Comprobado si existen nuevas incidencias")
        df_dict_info = DataFrame(category_dict_info)
        df_dict_info_length = df_dict_info.shape[0]
        if df_dict_info_length == 0:
            LOGGER.info(
                "No se va a guardar el diccionario de links recorridos. Razón: No han aparecido nuevas incidencias",
            )
        else:
            if self._df_dict_category is None:
                LOGGER.info(
                    "Creando un diccionario de datos con las nuevas incidencias encontradas"
                )
                self._df_dict_category = df_dict_info
            else:
                LOGGER.info(
                    "Actualizando el diccionario de datos con las nuevas incidencias encontradas"
                )
                self._df_dict_category = concat(
                    [self._df_dict_category, df_dict_info], axis=0, ignore_index=True
                )
            LOGGER.info(
                f"Cantidad de incidencias a ser guardadas: {df_dict_info_length}"
            )
            self._df_dict_category.sort_values(DATA_DICT_HEADERS[0], inplace=True)
            self._df_dict_category.to_csv(
                self._df_dict_category_filename,
                header=False,
                index=False,
                encoding="utf-8-sig",
            )
            LOGGER.info(
                f"Diccionarios de datos guardado satisfactoriamente con el nombre {self._df_dict_category_filename} en la ruta {ROOT_PATH}",
            )

        LOGGER.info("Filtrando categoría Especiales")
        category_info_link = {
            key: category_info_link[key]
            for key in category_info_link
            if key != "Especiales"
        }
        LOGGER.info("Categorías principales recuperadas satisfactoriamente\n")
        return DataFrame(
            {
                "Id_0": [
                    extract_text(r"/.*/(.*)/", x) for x in category_info_link.values()
                ],
                "Name_0": category_info_link.keys(),
                "Category_path_0": ["" for _ in range(len(category_info_link))],
            }
        )

    def send_request_api(self, id_cat, name_subcat, path_subcat):
        """Realiza una petición a la api usando el id, nombre y path de una categoría de Saga Falabella y retorna el id, nombre y path de todas sus subcategorías

        Args:
            id_cat (str): Id de la categoría a extraer su información
            name_subcat (str): Nombre de la categoría a extraer su información
            path_subcat (str): Path de la categoría a extraer su información

        Returns:
            list: Lista de subcategorías
        """
        subcategory_info = []
        try:
            # Realizando la petición a la api usando algunos parámetros necesarios
            response = get(
                API_URL.format(
                    path_subcat,
                    id_cat,
                    quote_plus(name_subcat),
                ),
                headers=API_HEADERS,
            )
            filters_value = response.json()["data"]["facets"][:4]
            # Recorriendo los 4 primeros filtros que posee la categoría
            for filter_value in filters_value[::-1]:
                # Comprobando si uno de los filtros contiene subcategorías
                if filter_value["name"] == "Categoría":
                    data_values = filter_value["values"]
                    # Guardando la información de una subcategoría
                    subcategory_info += [
                        [
                            id_cat,
                            item["id"],
                            item["title"],
                            item["url"].replace("+", "%20"),
                        ]
                        for item in data_values
                    ]
                    break
        except (KeyError, IndexError, JSONDecodeError):
            pass
        return subcategory_info

    def get_subcategory_info(self, column_values, whole_id):
        """Retorna un conjunto de datos que contiene toda la información de las subcategorías de saga falabella

        Args:
            column_values (list): Lista de valores a ser usados para la extracción de subcategorías
            whole_id (list): Lista de ids de todas las categorías que ha obtenido el scraper hasta el momento
        Returns:
            pandas.core.frame.DataFrame: Instancia de la clase DataFrame
        """
        LOGGER.info(
            f"Se van a extraer información de {len(column_values)} subcategorías"
        )
        subcategory_info = []
        # Realizando peticiones a la api
        futures_list_cat = [
            THREAD.submit(self.send_request_api, *category_level)
            for category_level in column_values
        ]

        # Recorriendo y registrando el resultado de todas las peticiones a la api
        for futures in as_completed(futures_list_cat):
            subcategory_info += futures.result()

        LOGGER.info(f"Se han extraído {len(subcategory_info)} subcategoría(s) nueva(s)")
        df_subcat_info = DataFrame(
            subcategory_info,
            columns=["Id", "Id_subcat", "Subcategory", "Category_path"],
        )

        LOGGER.info("Comprobando que no existan duplicados")
        index_values = [
            id_index
            for id_index, id_subcat in enumerate(df_subcat_info["Id_subcat"].values)
            if id_subcat in whole_id
        ]
        if len(index_values) > 0:
            LOGGER.info(f"Cantidad de subcategorías repetidas: {len(index_values)}")
            df_subcat_info.drop(index_values, inplace=True)
        return df_subcat_info

    def extract_categories(self, level):
        """Extrae la información de las categorías de saga falabella hasta cierto nivel de profundidad

        Args:
            level (int): Profundidad del árbol de categorías de saga falabella
        """
        LOGGER.info(
            f"Extrayendo el árbol de categorías de saga falabella con profundidad {level}",
        )
        if level <= 0:
            LOGGER.error(
                f"La cantidad de niveles de jerarquía de la clasificación de las categorías debe ser mayor o igual a 0",
            )
            self._driver.quit()
            return

        LOGGER.info("Entrando a la página web de la tienda de saga falabella")
        self._driver.get(URL_FALABELLA)
        self._driver.get(URL_FALABELLA) # Truco para eliminar varias ventanas molestosas

        LOGGER.info("Cerrando ventanas emergentes")
        self.close_popups()

        self._df_category = self.get_category_info(self.get_menu_links())
        self._df_category.sort_values("Id_0", inplace=True)
        df_subcategory = self._df_category
        whole_id = self._df_category["Id_0"].values.tolist()

        LOGGER.info("Cerrando navegador web")
        self._driver.quit()
        if level == 1:
            LOGGER.info(
                f"Se ha especificado nivel de profundidad {level}. No se va a extraer la información de las subcategorías.",
            )
            return

        LOGGER.info("Extrayendo información de las subcategorías")
        for i in range(1, level):
            LOGGER.info(f"Obteniendo información de las subcategorías de nivel {i}")
            id_prev = "Id_" + str(i - 1)
            name_prev = "Name_" + str(i - 1)
            category_path_prev = "Category_path_" + str(i - 1)
            df_subcategory = self.get_subcategory_info(
                df_subcategory[
                    [id_prev, name_prev, category_path_prev]
                ].values.tolist(),
                whole_id,
            )

            # Comprobando si ya no hay más resultados
            if df_subcategory.shape[0] == 0:
                level = i
                LOGGER.info(
                    f"Se ha llegado al máximo de profundidad con un valor de {level}.",
                )
                break

            df_subcategory.rename(
                {
                    "Id": id_prev,
                    "Subcategory": "Name_" + str(i),
                    "Id_subcat": "Id_" + str(i),
                    "Category_path": "Category_path_" + str(i),
                },
                axis=1,
                inplace=True,
            )
            df_subcategory.drop_duplicates(
                ["Id_" + str(i)], keep="first", inplace=True, ignore_index=True
            )
            whole_id += df_subcategory["Id_" + str(i)].values.tolist()

            LOGGER.info("Agregando las subcategorías encontradas al DataFrame original")
            self._df_category = self._df_category.merge(
                df_subcategory, how="left", left_on=id_prev, right_on=id_prev
            )
            LOGGER.info(f"Subcategorías de nivel {i} recuperadas satisfactoriamente")

        self._df_category.drop("Category_path_0", axis=1, inplace=True)
        self._df_category.replace("\n", "")
        LOGGER.info(
            f"Extracción de las categorías con un nivel de profundidad {level} completado satisfactoriamente\n",
        )

    def save_data(self, folder, filename, encoding="utf-8-sig"):
        """Guarda los datos o errores obtenidos durante la ejecución del scraper

        Args:
            folder (str): Ruta del archivo
            filename (str): Nombre del archivo
            encoding (str): Codificación usada para guardar el archivo. Defaults to "utf-8-sig"
        """
        LOGGER.info("Guardando la data")
        quantity = self._df_category.shape[0]
        self._metadata.quantity = quantity

        # Comprobando que el dataset contenga información
        if quantity == 0:
            LOGGER.info(
                f"El archivo de datos no se va a guardar por no tener información",
            )
            return

        # Generando la ruta donde se va a guardar la información
        datetime_obj = datetime.strptime(self._metadata.execution_date, "%d/%m/%Y")
        filepath = path.join(folder, datetime_obj.strftime("%d-%m-%Y"))
        filename = (
            filename
            + "_"
            + datetime_obj.strftime("%d%m%Y")
            + "_"
            + str(quantity)
            + ".csv"
        )

        # Verificando si la ruta donde se va a guardar la información existe
        if not path.exists(filepath):
            # Creando la ruta donde se va a guardar la información
            makedirs(filepath)
        self._df_category.to_csv(
            path.join(filepath, filename),
            sep=";",
            index=False,
            encoding=encoding,
        )
        LOGGER.info(
            f"El archivo de datos {filename} ha sido guardado correctamente en la ruta {path.join(ROOT_PATH, filepath)}",
        )

    def save_metadata(self, filename, sheet_name):
        """Guarda la información de la metadata generada durante la ejecución del scraper

        Args:
            filename (str): Nombre del archivo
            sheet_name (str): Nombre de la hoja de cálculo
        """
        # Guardando los parametros finales del tiempo de ejecución del scraper
        self._metadata.set_param_final()
        LOGGER.info("Guardando la metadata")
        # Variable que indica si el encabezados existe o no en el archivo de excel
        header_exist = False

        # Verificando si el archivo existe o no
        if path.isfile(filename):
            wb_time = load_workbook(filename)
            # Comprobando si ya existe un sheet con el nombre indicado en la variable sheet_name
            if sheet_name not in [ws.title for ws in wb_time.worksheets]:
                # Creando un nuevo sheet
                wb_time.create_sheet(sheet_name)
            else:
                header_exist = True
        else:
            wb_time = Workbook()
            wb_time.worksheets[0].title = sheet_name

        # Seleccionar el sheet deseado donde se va a guardar la información
        worksheet = wb_time[sheet_name]

        # Comprobando si el encabezados existen o no
        if not header_exist:
            keys = [
                "Fecha",
                "Hora Inicio",
                "Hora Fin",
                "Cantidad",
                "Tiempo Ejecucion (min)",
                "Categorias / Minuto",
                "Errores",
            ]
            worksheet.append(keys)

        values = list(self._metadata.__dict__.values())[1:]
        worksheet.append(values)
        wb_time.save(filename)
        wb_time.close()
        LOGGER.info(
            f"El archivo de la metadata del scraper {filename} ha sido guardado correctamente en la ruta {ROOT_PATH}",
        )


def config_log(log_folder, log_filename, log_file_mode="w", log_file_encoding="utf-8"):
    """Función que configura los logs para rastrear al programa

    Args:
        log_folder (str): Carpeta donde se va a generar el archivo log
        log_filename (str): Nombre del archivo log a ser generado
        log_file_mode (str, optional): Modo de guardado del archivo. Defaults to "w".
        log_file_encoding (str, optional): Codificación usada para el archivo. Defaults to "utf-8".
    """
    # Generando la ruta donde se va a guardar los registros de ejecución
    log_path = path.join(log_folder, CURRENT_DATE.strftime("%d-%m-%Y"))
    log_filename = log_filename + "_" + CURRENT_DATE.strftime("%d%m%Y") + ".log"

    # Verificando si la ruta donde se va a guardar los registros de ejecución existe
    if not path.exists(log_path):
        makedirs(log_path)

    # Agregando los handlers al logger
    formatter = Formatter("%(asctime)s - %(levelname)s - %(message)s")
    stream_handler = StreamHandler(stdout)
    stream_handler.setFormatter(formatter)
    file_handler = FileHandler(
        path.join(log_path, log_filename), log_file_mode, log_file_encoding
    )
    file_handler.setFormatter(formatter)
    LOGGER.handlers = [stream_handler, file_handler]
    LOGGER.propagate = False
    LOGGER.setLevel(INFO)


def validate_params(parameters):
    """Función que valida si los parámetros a usar están definidos

    Args:
        parameters (list): Lista de parámetros

    Returns:
        bool: Booleano que indica si los parámetros están definidos o no
    """
    for param in parameters:
        if not param or param == "":
            return False
    return True


def extract_text(pattern, text, n=1):
    """Extrae el texto deseado de una cadena dado una expresión regular

    Args:
        pattern (str): Expresión regular a utilizar para la búsqueda
        text (str): Texto donde se va a realizar la búsqueda
        n (int, optional): Número de grupo a recuperar. Defaults to 1.

    Returns:
        str or None: Texto encontrado o vacío
    """
    groups = search(pattern, text)
    return groups.group(n) if groups else groups


def main():
    try:
        # Formato para el debugger
        config_log("Log", "fb_ropa_log")
        LOGGER.info("Validando parámetros a usar")
        if not validate_params(
            [
                URL_FALABELLA,
                DATA_FILENAME,
                DATA_FOLDER,
                DATA_DICT_FILENAME,
                METADATA_FILENAME,
                METADATA_SHEET_NAME,
                API_URL,
            ]
        ):
            LOGGER.error("Parámetros incorrectos")
            return
        LOGGER.info("Parámetros válidos")

        LOGGER.info("Inicializando scraper")
        scraper = ScraperFalabellaCategory(DATA_DICT_FILENAME)
        LOGGER.info("Scraper inicializado satisfactoriamente")

        LOGGER.info("Extrayendo las categorias de falabella")
        scraper.extract_categories(5)

        LOGGER.info("Guardando toda la información generada por el scraper")
        scraper.save_data(DATA_FOLDER, DATA_FILENAME)
        scraper.save_metadata(METADATA_FILENAME, METADATA_SHEET_NAME)
        LOGGER.info("Programa finalizado")

    except Exception as error:
        Error(error).imprimir_error()
        LOGGER.error("Programa ejecutado con fallos")
        del scraper # Forzando el cierre del navegador web si es que se requiere
    finally:
        # Liberar el archivo log
        shutdown()


if __name__ == "__main__":
    main()
